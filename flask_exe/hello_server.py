from flask import Flask, render_template, request, redirect, url_for, session
import random
import datetime
import os
try:
    import openpyxl
except Exception:
    openpyxl = None

app = Flask(__name__)
app.secret_key = 'change-this-secret'

@app.route('/')
def index():
    dt_now = datetime.datetime.now()
    return render_template('htmls/login.html', time=dt_now.strftime('%Y年%m月%d日 %H:%M:%S'))

@app.route('/password')
def password():
    login = request.args.get('login')
    forgot = str(random.randint(0, 4))  # 0〜4の乱数
    return render_template('htmls/password.html', login=login, forgot=forgot)


def read_all_reports():
    """Read all reports from Excel file."""
    reports = []
    if not openpyxl:
        return reports
    path = os.path.join(app.root_path, 'data', 'reports.xlsx')
    if not os.path.exists(path):
        return reports
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            reports.append({
                'row_id': idx,
                'username': row[0],
                'date': row[1],
                'title': row[2],
                'content': row[3],
                'datetime': row[4],
                'comment': row[5] if len(row) > 5 else None,
                'comment_teacher': row[6] if len(row) > 6 else None
            })
    except Exception:
        pass
    return reports


def read_user_reports(username):
    """Read reports for a specific user."""
    reports = []
    if not openpyxl:
        return reports
    path = os.path.join(app.root_path, 'data', 'reports.xlsx')
    if not os.path.exists(path):
        return reports
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            if row[0] == username:
                reports.append({
                    'row_id': idx,
                    'username': row[0],
                    'date': row[1],
                    'title': row[2],
                    'content': row[3],
                    'datetime': row[4],
                    'comment': row[5] if len(row) > 5 else None,
                    'comment_teacher': row[6] if len(row) > 6 else None
                })
    except Exception:
        pass
    return reports


def save_report(username, date, title, content):
    """Save a report to Excel file."""
    if not openpyxl:
        return False
    path = os.path.join(app.root_path, 'data', 'reports.xlsx')
    dt_now = datetime.datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')
    
    try:
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['ユーザー名', '日付', 'タイトル', '内容', '投稿日時', 'コメント', 'コメント教員'])
        
        ws.append([username, date, title, content, dt_now, ''])
        wb.save(path)
        return True
    except Exception as e:
        print(f"Error saving report: {e}")
        return False


def save_comment(row_id, comment, teacher_name=None):
    """Save a teacher's comment to a report."""
    if not openpyxl:
        return False
    path = os.path.join(app.root_path, 'data', 'reports.xlsx')
    
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # row_id is the Excel row number (2-indexed for data rows)
        ws.cell(row=row_id, column=6).value = comment
        if teacher_name:
            ws.cell(row=row_id, column=7).value = teacher_name
        wb.save(path)
        return True
    except Exception as e:
        print(f"Error saving comment: {e}")
        return False


def read_users_from_excel(path=None):
    """Return dict of username->{'password': str, 'role': str} from an Excel file if available."""
    users = {}
    if path is None:
        path = os.path.join(app.root_path, 'data', 'users.xlsx')
    if openpyxl and os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            username = str(row[0]).strip() if row[0] is not None else ''
            password = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            role = str(row[2]).strip() if len(row) > 2 and row[2] is not None else 'student'
            if username:
                users[username] = {'password': password, 'role': role}
        return users
 


@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username', '')
    password = request.form.get('password', '')
    users = read_users_from_excel()
    # authenticate
    if username in users and users[username]['password'] == password:
        session['username'] = username
        session['role'] = users[username]['role']
        return redirect(url_for('home'))
    # login failed
    return render_template('htmls/login.html', time=datetime.datetime.now().strftime('%Y年%m月%d日 %H:%M:%S'), error='ユーザー名かパスワードが違います')


@app.route('/home')
def home():
    username = session.get('username')
    if not username:
        return redirect(url_for('index'))
    role = session.get('role', 'student')
    dt_now = datetime.datetime.now()
    time = dt_now.strftime('%Y年%m月%d日 %H:%M:%S')
    
    if role == 'teacher':
        # 教員用：全生徒の日誌を表示
        all_reports = read_all_reports()
        search_query = request.args.get('search', '').strip()
        
        # 検索クエリがある場合はフィルタリング
        if search_query:
            filtered_reports = []
            search_lower = search_query.lower()
            for report in all_reports:
                # ユーザー名、タイトル、内容、コメントから検索
                if (search_lower in str(report.get('username', '')).lower() or
                    search_lower in str(report.get('title', '')).lower() or
                    search_lower in str(report.get('content', '')).lower() or
                    search_lower in str(report.get('comment', '')).lower() or
                    search_lower in str(report.get('date', '')).lower()):
                    filtered_reports.append(report)
            all_reports = filtered_reports
        
        return render_template('htmls/teacher_home.html', username=username, time=time, all_reports=all_reports, search_query=search_query)
    else:
        # 生徒用：日誌入力フォームと自分の日誌を表示
        my_reports = read_user_reports(username)
        return render_template('htmls/student_home.html', username=username, time=time, reports=my_reports)


@app.route('/submit_report', methods=['POST'])
def submit_report():
    username = session.get('username')
    if not username:
        return redirect(url_for('index'))
    date = request.form.get('date', '').strip()
    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()
    if date and title and content:
        save_report(username, date, title, content)
    return redirect(url_for('home'))


@app.route('/submit_comment', methods=['POST'])
def submit_comment():
    role = session.get('role', 'student')
    if role != 'teacher':
        return redirect(url_for('home'))
    
    teacher_name = session.get('username')
    row_id = request.form.get('row_id', '').strip()
    comment = request.form.get('comment', '').strip()
    
    if row_id and comment:
        try:
            row_id = int(row_id)
            save_comment(row_id, comment, teacher_name)
        except ValueError:
            pass
    
    return redirect(url_for('home'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(debug=True)
